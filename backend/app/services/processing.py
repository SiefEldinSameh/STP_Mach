"""
Async job management - runs heavy inference in background threads.
"""

import asyncio
import copy
import csv
import io
import json
import logging
import os
import threading
import uuid
import zipfile
from concurrent.futures import Future, ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Literal, Optional

from openpyxl import Workbook

from app.config import OUTPUT_DIR, PROCESSING_MAX_WORKERS, SUBSCRIBE_URL
from app.pipeline.inference import run_pipeline
from app.services.health import health_tracker

logger = logging.getLogger(__name__)

_executor = ThreadPoolExecutor(
    max_workers=max(1, PROCESSING_MAX_WORKERS),
    thread_name_prefix="table-extract",
)
_jobs: Dict[str, dict] = {}
_job_futures: Dict[str, Future] = {}
_jobs_lock = threading.Lock()


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _set_job_fields(job_id: str, **fields):
    with _jobs_lock:
        job = _jobs.get(job_id)
        if job is not None:
            job.update(fields)


def _get_job_field(job_id: str, field: str):
    with _jobs_lock:
        job = _jobs.get(job_id)
        if job is None:
            return None
        return job.get(field)


def _snapshot_job_in_memory(job_id: str) -> Optional[dict]:
    with _jobs_lock:
        job = _jobs.get(job_id)
        if job is None:
            return None
        return copy.deepcopy(job)


def _register_job_future(job_id: str, future: Future) -> None:
    with _jobs_lock:
        _job_futures[job_id] = future


def _cleanup_finished_futures() -> None:
    with _jobs_lock:
        finished = [job_id for job_id, future in _job_futures.items() if future.done()]
        for job_id in finished:
            _job_futures.pop(job_id, None)


def _has_tables(result: Optional[dict]) -> bool:
    if not result:
        return False
    return any(page.get("tables") for page in result.get("pages", []))


def _run_inference_sync(job_id: str, file_path: str, original_filename: str):
    """Synchronous wrapper for the pipeline - runs in a worker thread."""
    try:
        _set_job_fields(job_id, progress_stage="loading")
        output_dir = _get_job_field(job_id, "output_dir")
        if not output_dir:
            raise RuntimeError(f"Job {job_id} is missing an output directory")
        result = run_pipeline(
            file_path,
            output_dir=output_dir,
            display_filename=original_filename,
            stage_callback=lambda stage: _set_job_fields(job_id, progress_stage=stage),
        )
        _set_job_fields(
            job_id,
            status="completed",
            result=result,
            filename=result["filename"],
            total_latency_ms=result["total_latency_ms"],
            finished_at=_utc_now(),
            progress_stage="completed",
            stage_timings_ms=result.get("stage_timings_ms", {}),
        )
        completed_job = _snapshot_job_in_memory(job_id)
        _save_job_to_disk(job_id, completed_job)
        health_tracker.record_request(
            result["total_latency_ms"],
            True,
            stage_timings=result.get("stage_timings_ms", {}),
            job_summary={
                "job_id": job_id,
                "filename": result["filename"],
                "status": "completed",
                "total_latency_ms": result["total_latency_ms"],
                "started_at": completed_job["started_at"],
                "finished_at": completed_job["finished_at"],
                "progress_stage": completed_job["progress_stage"],
                "error": None,
            },
        )
    except Exception as exc:
        logger.exception("Job %s failed", job_id)
        _set_job_fields(
            job_id,
            status="error",
            error=str(exc),
            finished_at=_utc_now(),
            progress_stage="error",
        )
        failed_job = _snapshot_job_in_memory(job_id) or {}
        _save_job_to_disk(job_id, failed_job)
        health_tracker.record_request(
            0,
            False,
            stage_timings=failed_job.get("stage_timings_ms"),
            job_summary={
                "job_id": job_id,
                "filename": original_filename,
                "status": "error",
                "total_latency_ms": 0,
                "started_at": failed_job.get("started_at"),
                "finished_at": failed_job.get("finished_at"),
                "progress_stage": failed_job.get("progress_stage"),
                "error": str(exc),
            },
        )


async def submit_job(file_path: str, original_filename: str) -> str:
    """Submit a file for async processing. Returns job_id."""
    _cleanup_finished_futures()
    job_id = str(uuid.uuid4())
    output_dir = os.path.join(OUTPUT_DIR, job_id)
    os.makedirs(output_dir, exist_ok=True)

    with _jobs_lock:
        _jobs[job_id] = {
            "status": "processing",
            "file_path": file_path,
            "result": None,
            "error": None,
            "filename": original_filename,
            "total_latency_ms": None,
            "started_at": _utc_now(),
            "finished_at": None,
            "progress_stage": "queued",
            "output_dir": output_dir,
            "stage_timings_ms": {},
        }

    loop = asyncio.get_running_loop()
    future = loop.run_in_executor(_executor, _run_inference_sync, job_id, file_path, original_filename)
    _register_job_future(job_id, future)
    return job_id


def _save_job_to_disk(job_id: str, job: dict) -> None:
    """Persist a completed job to disk so it survives server restarts."""
    output_dir = job.get("output_dir")
    if not output_dir:
        return
    result_path = Path(output_dir) / "result.json"
    try:
        payload = {
            "job_id": job_id,
            "status": job.get("status"),
            "filename": job.get("filename"),
            "total_latency_ms": job.get("total_latency_ms"),
            "started_at": job.get("started_at"),
            "finished_at": job.get("finished_at"),
            "progress_stage": job.get("progress_stage"),
            "stage_timings_ms": job.get("stage_timings_ms", {}),
            "error": job.get("error"),
            "result": job.get("result"),
        }
        result_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    except Exception:
        logger.warning("Failed to persist result.json for job %s", job_id)


def _load_job_from_disk(job_id: str) -> Optional[dict]:
    """Load a previously completed job from disk (fallback when not in memory)."""
    result_path = Path(OUTPUT_DIR) / job_id / "result.json"
    if not result_path.is_file():
        return None
    try:
        data = json.loads(result_path.read_text(encoding="utf-8"))
        return {
            "status": data.get("status", "completed"),
            "filename": data.get("filename"),
            "total_latency_ms": data.get("total_latency_ms"),
            "started_at": data.get("started_at"),
            "finished_at": data.get("finished_at"),
            "progress_stage": data.get("progress_stage", "completed"),
            "stage_timings_ms": data.get("stage_timings_ms", {}),
            "error": data.get("error"),
            "result": data.get("result"),
            "output_dir": str(Path(OUTPUT_DIR) / job_id),
            "file_path": None,
        }
    except Exception:
        logger.warning("Failed to load result.json for job %s", job_id)
        return None


def get_job(job_id: str) -> Optional[dict]:
    """Get job by ID — checks memory first, then falls back to disk."""
    job = _snapshot_job_in_memory(job_id)
    if job is not None:
        return job
    return _load_job_from_disk(job_id)


def get_processing_runtime() -> dict[str, int]:
    """Return thread-pool activity for health/debugging."""
    _cleanup_finished_futures()
    with _jobs_lock:
        processing_jobs = [job for job in _jobs.values() if job.get("status") == "processing"]
        queued = sum(1 for job in processing_jobs if job.get("progress_stage") == "queued")
        active = max(0, len(processing_jobs) - queued)
        tracked = len(_job_futures)

    return {
        "max_workers": _executor._max_workers,
        "active_jobs": active,
        "queued_jobs": queued,
        "tracked_jobs": tracked,
    }


def shutdown_processing_executor() -> None:
    _executor.shutdown(wait=False, cancel_futures=False)


def update_job_cell(job_id: str, page: int, table_id: int, row: int, col: int, text: str) -> bool:
    """Update a cell's text in the job results (in-memory or disk-backed)."""
    job = get_job(job_id)
    if not job or job["status"] != "completed" or not job["result"]:
        return False

    with _jobs_lock:
        from_disk = job_id not in _jobs

    for page_data in job["result"].get("pages", []):
        if page_data["page"] != page:
            continue
        for table in page_data["tables"]:
            if table["table_id"] != table_id:
                continue
            for cell in table["cells"]:
                if cell["row"] == row and cell["col"] == col:
                    cell["text"] = text
                    if from_disk:
                        _save_job_to_disk(job_id, job)
                    return True
    return False


def job_has_tables(job_id: str) -> bool:
    job = get_job(job_id)
    if not job or job["status"] != "completed":
        return False
    return _has_tables(job.get("result"))


def _table_to_matrix(table: dict) -> list[list[str]]:
    cells = table.get("cells", [])
    if not cells:
        return []

    n_rows = max(cell["row"] + cell["rowspan"] for cell in cells)
    n_cols = max(cell["col"] + cell["colspan"] for cell in cells)
    matrix = [["" for _ in range(n_cols)] for _ in range(n_rows)]

    for cell in cells:
        row = cell["row"]
        col = cell["col"]
        matrix[row][col] = cell.get("text", "")
        for row_offset in range(cell["rowspan"]):
            for col_offset in range(cell["colspan"]):
                if row_offset == 0 and col_offset == 0:
                    continue
                matrix[row + row_offset][col + col_offset] = ""

    return matrix


def generate_csv_for_result(result: dict, csv_format: str = "matrix") -> Optional[str]:
    """Generate CSV content from structured results."""
    if not _has_tables(result):
        return None

    output = io.StringIO()
    writer = csv.writer(output)

    if csv_format == "cells":
        writer.writerow(["page", "table_id", "row", "col", "rowspan", "colspan", "text"])
        for page_data in result.get("pages", []):
            for table in page_data.get("tables", []):
                for cell in table.get("cells", []):
                    writer.writerow(
                        [
                            page_data["page"],
                            table["table_id"],
                            cell["row"],
                            cell["col"],
                            cell["rowspan"],
                            cell["colspan"],
                            cell["text"],
                        ]
                    )
        return output.getvalue()

    for page_data in result.get("pages", []):
        for table in page_data.get("tables", []):
            writer.writerow([f"Page {page_data['page'] + 1}", f"Table {table['table_id'] + 1}"])
            for row in _table_to_matrix(table):
                writer.writerow(row)
            writer.writerow([])

    return output.getvalue()


def _export_promo_text() -> str:
    msg = (
        "If you want faster or up to 10× faster processing, subscribe with us for marketing updates."
    )
    if SUBSCRIBE_URL:
        return f"{msg} {SUBSCRIBE_URL}"
    return msg


def _safe_excel_sheet_title(raw: str, max_len: int = 31) -> str:
    invalid = r"[]*?:\/\\"
    s = "".join("_" if c in invalid else c for c in raw).strip() or "Sheet"
    return s[:max_len]


def generate_xlsx_for_result(
    result: dict, xlsx_format: Literal["matrix", "cells"] = "matrix"
) -> Optional[bytes]:
    if not _has_tables(result):
        return None

    wb = Workbook()

    if xlsx_format == "cells":
        ws = wb.active
        ws.title = "Cells"
        ws.append(["page", "table_id", "row", "col", "rowspan", "colspan", "text"])
        for page_data in result.get("pages", []):
            for table in page_data.get("tables", []):
                for cell in table.get("cells", []):
                    ws.append(
                        [
                            page_data["page"],
                            table["table_id"],
                            cell["row"],
                            cell["col"],
                            cell["rowspan"],
                            cell["colspan"],
                            cell.get("text", ""),
                        ]
                    )
    else:
        table_index = 0
        for page_data in result.get("pages", []):
            for table in page_data.get("tables", []):
                title = _safe_excel_sheet_title(f"Table {table_index}")
                if table_index == 0:
                    ws = wb.active
                    ws.title = title
                else:
                    ws = wb.create_sheet(title)
                for cell in table.get("cells", []):
                    r = cell["row"] + 1
                    c = cell["col"] + 1
                    rs = max(1, int(cell.get("rowspan", 1)))
                    cs = max(1, int(cell.get("colspan", 1)))
                    ws.cell(row=r, column=c, value=cell.get("text", ""))
                    if rs > 1 or cs > 1:
                        ws.merge_cells(
                            start_row=r,
                            start_column=c,
                            end_row=r + rs - 1,
                            end_column=c + cs - 1,
                        )
                table_index += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_crops_zip(job_id: str) -> Optional[io.BytesIO]:
    job = get_job(job_id)
    if not job or job["status"] != "completed":
        return None

    crops_dir = Path(job["output_dir"]) / "crops"
    crop_files = sorted(crops_dir.glob("*.png"))
    if not crop_files:
        return None

    archive = io.BytesIO()
    with zipfile.ZipFile(archive, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        for crop_file in crop_files:
            zip_file.write(crop_file, arcname=crop_file.name)
    archive.seek(0)
    return archive
